"""有货未展示看板 —— 纯本地数据层（不依赖任何网页/Flask）。

核心：用店面展示数据 (display) 比对 non-discontinue 且有货的产品 (stock)，
按【店面】逐个产品给出：有没有货、在该店面有没有展示、图片。

数据来源（按优先级）
1. 环境变量 INSTOCK_STOCK_CSV / INSTOCK_DISPLAY_CSV
2. INSTOCK_DATA_DIR 目录下的 stock.csv / display.csv
3. INSTOCK_REGION（如 NZ/CA）→ Output-{region}/ 下的 stock/display 文件

支持 CSV 和 Excel（.xlsx）；列名不需要完全一致，_pick 会做模糊匹配。
"""

import csv
import json
import os
import re
from pathlib import Path
from urllib.parse import quote, urlsplit, urlunsplit

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

ROOT_DIR = Path(__file__).parent
RUNNER_CONFIG_FILE = ROOT_DIR / "region_runner_config.json"
EXEMPTION_CONFIG_FILE = ROOT_DIR / "family_exemption.json"
IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".gif")

CODE_KEYS = ["productcode", "product_code", "sku", "itemcode", "item_code",
             "code", "productid", "product_id", "product", "item"]
BLACKLIST_STEMS = ["blacklist", "sku_blacklist", "black_list", "product_blacklist"]
BLACKLIST_KEYS = ["sku", "blacklist", "blacklistsku"] + CODE_KEYS
NAME_KEYS = ["productname", "product_name", "name", "description", "desc", "title"]
FAMILY_KEYS = ["family", "productfamily", "product_family", "category",
               "categoryname", "category_name", "group", "producttype", "type",
               "productfamilyname"]
STOCK_KEYS = ["stockqty", "stock_qty", "stock", "qty", "quantity", "onhand",
              "on_hand", "onhandqty", "on_hand_qty", "qtyonhand", "qty_on_hand",
              "available", "availableqty", "available_qty", "availablestock",
              "soh", "totalstock", "total_stock"]
PRICE_KEYS = ["price", "unitprice", "unit_price", "sellprice", "sell_price",
              "saleprice", "sale_price", "retailprice", "retail_price"]
DISCONTINUE_KEYS = ["discontinued", "isdiscontinued", "is_discontinued", "isdisconti",
                    "discontinue", "discontinueflag", "discontinue_flag", "status"]
STORE_KEYS = ["store", "storename", "store_name", "warehouse", "warehousename",
              "warehouse_name", "location", "branch", "shop", "displaywarehouse",
              "display_warehouse", "site"]
IMAGE_KEYS = ["imagefile", "image", "imageurl", "image_url", "img",
              "picture", "photo", "thumbnail", "thumb"]

ALL_STORES = "全部店面"
UNCATEGORIZED_FAMILIES = frozenset({"", "未分类", "UNCATEGORIZED", "N/A", "NONE", "未知"})

# 店面展示名 → 库存列（stock.xlsx 多仓交叉读取）
WAREHOUSE_LABELS = {
    "carbine": "Carbine",
    "walls": "Walls",
    "geraldconnelly": "GC",
    "northisland": "North Island",
}
REGION_STORE_STOCK_RULES = {
    "NZ": [
        (("onehunga", "westgate", "hamilton", "sleeplab"), ("carbine", "walls")),
        (("chch", "christchurch", "gerald", "treffers", "presale"), ("geraldconnelly",)),
    ],
}
DEFAULT_ALL_WAREHOUSES = ("carbine", "walls", "geraldconnelly")

# region_key -> cached bundle (invalidated when file mtime changes)
_REGION_CACHE = {}
_STORE_VIEW_CACHE = {}
_EXEMPTION_CONFIG_CACHE = None


def _pick(row, keys):
    lower = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        if key in lower:
            value = lower[key]
            if value is not None and str(value).strip() != "":
                return value
    return None


def _to_float(value):
    if value is None:
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except Exception:
        return None


def _is_discontinued(value):
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        try:
            return float(value) != 0
        except (TypeError, ValueError):
            pass
    text = str(value).strip().lower()
    if text in ("", "0", "0.0", "false", "no", "n", "active", "live"):
        return False
    if text in ("1", "1.0", "true", "yes", "y"):
        return True
    return "discontinue" in text


def _norm_code(value):
    return str(value).strip().upper() if value is not None else ""


def _norm_col_key(name):
    return re.sub(r"[^a-z0-9]", "", str(name).strip().lower())


def _classify_warehouse_column(col_name):
    """把 stock.xlsx 列名映射到标准仓名（CarbineSt / WallsStoc / GeraldConnellyStock 等）。"""
    key = _norm_col_key(col_name)
    if not key:
        return None
    if "carbine" in key:
        return "carbine"
    if "walls" in key:
        return "walls"
    if "geraldconnelly" in key or key.startswith("gc"):
        return "geraldconnelly"
    if "northisland" in key or key.startswith("northislar"):
        return "northisland"
    return None


def _extract_warehouse_stock(row):
    """从一行 stock 数据提取各仓库存数量。"""
    stock = {}
    for col, value in row.items():
        wh = _classify_warehouse_column(col)
        if not wh:
            continue
        qty = _to_float(value) or 0.0
        stock[wh] = stock.get(wh, 0.0) + qty
    return stock


def _warehouses_for_store(store_name, region_key):
    """根据所选店面，决定用哪些仓库列计算有货数量。"""
    if store_name == ALL_STORES:
        return DEFAULT_ALL_WAREHOUSES

    text = str(store_name).strip().lower()
    for patterns, warehouses in REGION_STORE_STOCK_RULES.get(region_key.upper(), []):
        if any(p in text for p in patterns):
            return warehouses

    if any(x in text for x in ("auck", "onehunga", "westgate", "hamilton", "north")):
        return ("carbine", "walls")
    if any(x in text for x in ("chch", "christ", "gerald", "treffers", "south")):
        return ("geraldconnelly",)
    return DEFAULT_ALL_WAREHOUSES


def _qty_from_warehouses(warehouse_stock, warehouse_keys):
    return sum(float(warehouse_stock.get(k, 0) or 0) for k in warehouse_keys)


def _stock_breakdown(warehouse_stock, warehouse_keys):
    parts = []
    for key in warehouse_keys:
        qty = int(warehouse_stock.get(key, 0) or 0)
        if qty > 0:
            parts.append(f"{WAREHOUSE_LABELS.get(key, key)} {qty}")
    return " + ".join(parts) if parts else ""


def _apply_store_stock(product, store, region_key):
    warehouse_stock = product.get("warehouse_stock") or {}
    warehouse_keys = _warehouses_for_store(store, region_key)
    if warehouse_stock:
        qty = _qty_from_warehouses(warehouse_stock, warehouse_keys)
        breakdown = _stock_breakdown(warehouse_stock, warehouse_keys)
    else:
        qty = float(product.get("stock_qty") or 0)
        breakdown = ""
    item = dict(product)
    item["stock_qty"] = qty
    item["in_stock"] = qty > 0
    item["stock_warehouses"] = warehouse_keys
    item["stock_breakdown"] = breakdown
    return item


def _read_csv(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _read_xlsx(path):
    if load_workbook is None:
        raise RuntimeError("读取 Excel 需要 openpyxl，请运行：pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        wb.close()
        return []
    columns = [str(h).strip() if h is not None else "" for h in headers]
    out = []
    for row in rows_iter:
        if row is None:
            continue
        item = {}
        empty = True
        for idx, col in enumerate(columns):
            if not col:
                continue
            value = row[idx] if idx < len(row) else None
            if value is not None and str(value).strip() != "":
                empty = False
            item[col] = value
        if not empty:
            out.append(item)
    wb.close()
    return out


def _read_table(path):
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        return _read_xlsx(path)
    return _read_csv(path)


def _file_mtime(path):
    p = Path(path)
    return p.stat().st_mtime if p.is_file() else 0.0


def normalize_url(url):
    """把含中文/空格的 URL 编码成可请求的地址。"""
    text = str(url).strip()
    if not text.lower().startswith(("http://", "https://")):
        return text
    parts = urlsplit(text)
    path = quote(parts.path, safe="/:@")
    query = quote(parts.query, safe="=&?/:;+") if parts.query else parts.query
    return urlunsplit((parts.scheme, parts.netloc, path, query, parts.fragment))


def _find_data_file(directory, stems):
    """在目录里按候选文件名找 stock/display（优先 xlsx，其次 csv）。"""
    base = Path(directory)
    if not base.is_dir():
        return None
    for stem in stems:
        for ext in (".xlsx", ".csv"):
            candidate = base / f"{stem}{ext}"
            if candidate.is_file():
                return candidate
    return None


def _find_region_data_file(output_dir, stems):
    """在 Output 目录及其子目录（latest、时间戳文件夹）里找数据文件。"""
    base = Path(output_dir)
    if not base.is_dir():
        return None

    found = _find_data_file(base, stems)
    if found:
        return found

    latest = base / "latest"
    found = _find_data_file(latest, stems)
    if found:
        return found

    stamp_dirs = sorted(
        (
            d for d in base.iterdir()
            if d.is_dir() and re.fullmatch(r"\d{8}_\d{6}", d.name)
        ),
        reverse=True,
    )
    for sub in stamp_dirs:
        found = _find_data_file(sub, stems)
        if found:
            return found
    return None


def _region_stock_stems(region_key):
    rk = region_key.upper()
    return ["stock", f"{rk}_stock", "product_stock_price", f"{rk}_product_stock_price"]


def _region_display_stems(region_key):
    rk = region_key.upper()
    return ["display", "display_with_families", f"{rk}_display",
            "store_display", f"{rk}_store_display"]


def _load_runner_regions():
    if not RUNNER_CONFIG_FILE.exists():
        return {}
    try:
        with RUNNER_CONFIG_FILE.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("regions", {}) if isinstance(data, dict) else {}
    except Exception:
        return {}


def _region_output_dir(region_key):
    regions = _load_runner_regions()
    cfg = regions.get(region_key.upper(), {})
    output_dir = (cfg.get("output_dir") or f"Output-{region_key.upper()}").strip()
    path = Path(output_dir)
    if not path.is_absolute():
        path = ROOT_DIR / path
    return path.resolve()


def resolve_sources(region=None):
    stock_env = os.getenv("INSTOCK_STOCK_CSV")
    display_env = os.getenv("INSTOCK_DISPLAY_CSV")
    data_dir = os.getenv("INSTOCK_DATA_DIR")
    region_key = (region or os.getenv("INSTOCK_REGION") or "").strip().upper()

    if stock_env and display_env:
        stock_path = Path(stock_env)
        return stock_path, Path(display_env), "csv", stock_path.parent
    if data_dir:
        d = Path(data_dir)
        return d / "stock.csv", d / "display.csv", "csv", d
    if region_key:
        d = _region_output_dir(region_key)
        stock_path = _find_region_data_file(d, _region_stock_stems(region_key))
        display_path = _find_region_data_file(d, _region_display_stems(region_key))
        if stock_path is None:
            stock_path = d / "stock.xlsx"
        if display_path is None:
            display_path = d / "display.xlsx"
        return stock_path, display_path, f"region-{region_key}", d
    raise ValueError(
        "未指定地区。请在界面选择 NZ/AU/CA，或设置环境变量 INSTOCK_REGION。"
    )


def list_regions():
    options = []
    for key, cfg in _load_runner_regions().items():
        out_dir = _region_output_dir(key)
        has_data = _find_region_data_file(out_dir, _region_stock_stems(key)) is not None
        options.append({
            "key": key,
            "label": cfg.get("label", key),
            "has_latest": has_data,
        })
    return options


def default_region():
    """返回第一个已有导出数据的地区，否则返回配置里的第一个地区。"""
    regions = list_regions()
    for r in regions:
        if r.get("has_latest"):
            return r["key"]
    return regions[0]["key"] if regions else None


def clear_region_cache(region=None):
    if region is None:
        _REGION_CACHE.clear()
        _STORE_VIEW_CACHE.clear()
        return
    region_key = str(region).strip().upper()
    _REGION_CACHE.pop(region_key, None)
    for key in list(_STORE_VIEW_CACHE):
        if key[0] == region_key:
            _STORE_VIEW_CACHE.pop(key, None)


def _load_blacklist(path):
    """读取黑名单 SKU 集合（列名 sku / 编码 / ProductCode 等均可）。"""
    codes = set()
    if not path or not Path(path).is_file():
        return codes
    for row in _read_table(path):
        code = _pick(row, BLACKLIST_KEYS)
        if code:
            codes.add(_norm_code(code))
    return codes


def _resolve_blacklist_path(data_dir):
    return _find_region_data_file(data_dir, BLACKLIST_STEMS)


def _load_region_bundle(region, force=False):
    """读取并缓存某地区的 stock/display 原始表（切换店面时复用，避免重复读 Excel）。"""
    region_key = str(region).strip().upper()
    stock_path, display_path, source, data_dir = resolve_sources(region_key)
    stock_mtime = _file_mtime(stock_path)
    display_mtime = _file_mtime(display_path)
    blacklist_path = _resolve_blacklist_path(data_dir)
    blacklist_mtime = _file_mtime(blacklist_path) if blacklist_path else None

    cached = _REGION_CACHE.get(region_key)
    if (
        not force
        and cached
        and cached["stock_mtime"] == stock_mtime
        and cached["display_mtime"] == display_mtime
        and cached.get("blacklist_mtime") == blacklist_mtime
    ):
        return cached

    if not Path(stock_path).is_file():
        raise FileNotFoundError(stock_path)
    if not Path(display_path).is_file():
        raise FileNotFoundError(display_path)

    display_rows = _read_table(display_path)
    by_store, display_details = _load_display(display_rows)
    blacklist = _load_blacklist(blacklist_path)
    bundle = {
        "region": region_key,
        "stock_path": stock_path,
        "display_path": display_path,
        "blacklist_path": str(blacklist_path) if blacklist_path else None,
        "blacklist": blacklist,
        "source": source,
        "data_dir": data_dir,
        "stock_mtime": stock_mtime,
        "display_mtime": display_mtime,
        "blacklist_mtime": blacklist_mtime,
        "stock_rows": _load_stock(_read_table(stock_path), data_dir),
        "by_store": by_store,
        "display_details": display_details,
        "display_row_count": len(display_rows),
    }
    _REGION_CACHE[region_key] = bundle
    return bundle


def _resolve_image(raw, code, data_dir):
    """把图片列的值解析成可用路径/URL；也支持按产品编码自动找本地图。"""
    if raw:
        text = str(raw).strip()
        if text.lower().startswith(("http://", "https://")):
            return normalize_url(text)
        p = Path(text)
        if not p.is_absolute():
            p = ROOT_DIR / p
        if p.is_file():
            return str(p)

    images_dir = Path(data_dir) / "images"
    if images_dir.is_dir():
        for stem in (str(code).strip(), str(code).strip().upper(), str(code).strip().lower()):
            for ext in IMAGE_EXTS:
                candidate = images_dir / f"{stem}{ext}"
                if candidate.is_file():
                    return str(candidate)
    return None


def _load_stock(rows, data_dir):
    out = []
    for row in rows:
        code = _pick(row, CODE_KEYS)
        if not code:
            continue
        warehouse_stock = _extract_warehouse_stock(row)
        if warehouse_stock:
            qty = sum(warehouse_stock.values())
        else:
            qty = _to_float(_pick(row, STOCK_KEYS)) or 0.0
        out.append({
            "code": str(code).strip(),
            "name": str(_pick(row, NAME_KEYS) or "").strip(),
            "family": str(_pick(row, FAMILY_KEYS) or "未分类").strip(),
            "stock_qty": qty,
            "warehouse_stock": warehouse_stock,
            "price": _to_float(_pick(row, PRICE_KEYS)),
            "discontinued": _is_discontinued(_pick(row, DISCONTINUE_KEYS)),
            "in_stock": qty > 0,
            "image": _resolve_image(_pick(row, IMAGE_KEYS), code, data_dir),
        })
    return out


def _load_display(rows):
    """返回 (by_store, display_details)。

    by_store: {store_name: set(codes)}
    display_details: {store_name: {norm_code: {code, family, name}}}
    """
    by_store = {}
    display_details = {}
    for row in rows:
        code = _pick(row, CODE_KEYS)
        if not code:
            continue
        store = _pick(row, STORE_KEYS)
        store = str(store).strip() if store else "（未标注店面）"
        norm = _norm_code(code)
        by_store.setdefault(store, set()).add(norm)
        display_details.setdefault(store, {})[norm] = {
            "code": str(code).strip(),
            "family": str(_pick(row, FAMILY_KEYS) or "").strip(),
            "name": str(_pick(row, NAME_KEYS) or "").strip(),
        }
    return by_store, display_details


def _load_exemption_config():
    global _EXEMPTION_CONFIG_CACHE
    if _EXEMPTION_CONFIG_CACHE is not None:
        return _EXEMPTION_CONFIG_CACHE
    if not EXEMPTION_CONFIG_FILE.is_file():
        _EXEMPTION_CONFIG_CACHE = {"auto_by_family": True, "groups": []}
        return _EXEMPTION_CONFIG_CACHE
    try:
        with EXEMPTION_CONFIG_FILE.open("r", encoding="utf-8") as f:
            data = json.load(f)
        _EXEMPTION_CONFIG_CACHE = data if isinstance(data, dict) else {"auto_by_family": True, "groups": []}
    except Exception:
        _EXEMPTION_CONFIG_CACHE = {"auto_by_family": True, "groups": []}
    return _EXEMPTION_CONFIG_CACHE


def _norm_group_token(text):
    return re.sub(r"\s+", " ", str(text).strip().upper())


def _product_matches_group(product, group):
    fam = _norm_group_token(product.get("family") or "")
    name = str(product.get("name") or "").lower()
    for f in group.get("families") or []:
        if fam == _norm_group_token(f):
            return True
    for pat in group.get("name_patterns") or []:
        if str(pat).lower() in name:
            return True
    return False


def _is_uncategorized_family(family):
    text = str(family or "").strip()
    if not text:
        return True
    upper = text.upper()
    return text in UNCATEGORIZED_FAMILIES or upper in UNCATEGORIZED_FAMILIES


def _resolve_exemption_group(product, config):
    for group in config.get("groups") or []:
        if _product_matches_group(product, group):
            label = (group.get("label") or group.get("key") or "").strip()
            key = (group.get("key") or label or "").strip()
            if _is_uncategorized_family(label) or _is_uncategorized_family(key):
                return None, label or "未分类"
            return _norm_group_token(key), label or key
    if config.get("auto_by_family", True):
        fam = str(product.get("family") or "未分类").strip()
        if _is_uncategorized_family(fam):
            return None, fam or "未分类"
        return _norm_group_token(fam), fam
    return None, None


def _display_anchor_codes(members, displayed_details, config):
    """同组内在该店面已展示的 SKU（含 display 表有、stock 表无的款）。"""
    group_key = next((m.get("exemption_group") for m in members if m.get("exemption_group")), None)
    if not group_key:
        return []

    codes = []
    seen = set()
    for m in members:
        if not m.get("displayed"):
            continue
        code = m.get("code")
        norm = _norm_code(code)
        if norm and norm not in seen:
            seen.add(norm)
            codes.append(code)

    for norm, meta in (displayed_details or {}).items():
        if norm in seen:
            continue
        pseudo = {
            "code": meta.get("code") or norm,
            "family": meta.get("family") or "",
            "name": meta.get("name") or "",
        }
        gkey, _ = _resolve_exemption_group(pseudo, config)
        if gkey == group_key:
            seen.add(norm)
            codes.append(pseudo["code"])
    return codes


def _eligible_for_exemption(product):
    if product.get("displayed") or not product.get("in_stock"):
        return False
    if product.get("gap"):
        return True
    return bool(product.get("discontinued"))


def _apply_family_exemptions(products, displayed_details=None):
    """同 exemption 组内已有展示 SKU 时，豁免组内其他有货未展示 SKU。

    锚点条件：在所选店面「已展示」即可（含 display 表有展示、stock 表无库存的款），
    不要求锚点有货。停产但有货未展示的 SKU 也可被同组已展示款豁免。
    """
    config = _load_exemption_config()
    for p in products:
        p["exempted"] = False
        p["exemption_reason"] = ""
        gkey, glabel = _resolve_exemption_group(p, config)
        p["exemption_group"] = gkey
        p["exemption_group_label"] = glabel

    by_group = {}
    for p in products:
        gk = p.get("exemption_group")
        if gk:
            by_group.setdefault(gk, []).append(p)

    exempted_n = 0
    for members in by_group.values():
        anchor_codes = _display_anchor_codes(members, displayed_details, config)
        if not anchor_codes:
            continue
        reason = f"同组已展示：{anchor_codes[0]}"
        if len(anchor_codes) > 1:
            extra = ", ".join(anchor_codes[1:3])
            suffix = "…" if len(anchor_codes) > 3 else ""
            reason = f"同组已展示：{anchor_codes[0]}, {extra}{suffix}"
        for m in members:
            if not _eligible_for_exemption(m):
                continue
            m["exempted"] = True
            if m.get("gap"):
                m["gap"] = False
            m["exemption_reason"] = reason
            exempted_n += 1
    return exempted_n


def list_stores(region=None):
    bundle = _load_region_bundle(region or default_region())
    by_store = bundle["by_store"]
    return [ALL_STORES] + sorted(by_store.keys())


SKU_PREFIX_LEN = 3


def sku_prefix(code, length=SKU_PREFIX_LEN):
    """取 SKU/编码前 N 位作为分类前缀（如 107-381 → 107）。"""
    text = str(code or "").strip().upper()
    return text[:length] if text else "???"


def aggregate_by_sku_prefix(products, prefix_len=SKU_PREFIX_LEN, store_specific=True):
    """按 SKU 前三位汇总；产品数仅计在产（non-discontinue）SKU。"""
    buckets = {}
    for p in products:
        if p.get("discontinued"):
            continue
        key = sku_prefix(p.get("code"), prefix_len)
        buckets.setdefault(key, []).append(p)

    rows = []
    for prefix, items in buckets.items():
        total = len(items)
        in_stock = [i for i in items if i.get("in_stock")]
        in_stock_n = len(in_stock)
        displayed_is = [i for i in in_stock if i.get("displayed")]
        if store_specific:
            gap_items = [i for i in items if i.get("gap")]
            exempted = [i for i in items if i.get("exempted")]
        else:
            gap_items = []
            exempted = []
        rows.append({
            "prefix": prefix,
            "total": total,
            "in_stock_count": in_stock_n,
            "in_stock_rate": round(in_stock_n / total * 100, 1) if total else None,
            "displayed_in_stock": len(displayed_is),
            "display_coverage_rate": (
                round(len(displayed_is) / in_stock_n * 100, 1) if in_stock_n else None
            ),
            "gap_count": len(gap_items),
            "exempted_count": len(exempted),
        })
    rows.sort(key=lambda r: (-r["gap_count"], -(r["in_stock_rate"] or 0), r["prefix"]))
    return rows


def build_products(store=None, only_gap=False, include_discontinued=False, region=None,
                   force_refresh=False):
    """核心：按店面逐个产品计算 有货/展示 状态与汇总指标。"""
    region_key = region or default_region()
    if not region_key:
        raise ValueError("region_runner_config.json 中未配置任何地区。")

    bundle = _load_region_bundle(region_key, force=force_refresh)
    view_key = (
        region_key,
        store or ALL_STORES,
        bool(include_discontinued),
        bundle.get("stock_mtime"),
        bundle.get("display_mtime"),
        bundle.get("blacklist_mtime"),
    )
    if not force_refresh and view_key in _STORE_VIEW_CACHE:
        cached = _STORE_VIEW_CACHE[view_key]
        if only_gap:
            out = dict(cached)
            out["products"] = [p for p in cached["products"] if p["gap"]]
            return out
        return cached

    stock_rows = bundle["stock_rows"]
    by_store = bundle["by_store"]
    stock_path = bundle["stock_path"]
    display_path = bundle["display_path"]
    source = bundle["source"]
    blacklist = bundle.get("blacklist") or set()
    blacklist_path = bundle.get("blacklist_path")

    stores = [ALL_STORES] + sorted(by_store.keys())
    if store is None:
        store = ALL_STORES

    if store == ALL_STORES:
        displayed_codes = set().union(*by_store.values()) if by_store else set()
    else:
        displayed_codes = by_store.get(store, set())

    store_specific = store != ALL_STORES

    products = []
    for p in stock_rows:
        if _norm_code(p["code"]) in blacklist:
            continue
        if not include_discontinued and p.get("discontinued"):
            continue
        item = _apply_store_stock(p, store, region_key)
        displayed = _norm_code(item["code"]) in displayed_codes
        if store_specific:
            gap = item["in_stock"] and (not item["discontinued"]) and (not displayed)
        else:
            gap = False
        item["displayed"] = displayed
        item["gap"] = gap
        products.append(item)

    if store_specific:
        store_display_details = bundle.get("display_details", {}).get(store, {})
        exempted_count = _apply_family_exemptions(products, store_display_details)
    else:
        config = _load_exemption_config()
        for p in products:
            p["exempted"] = False
            p["exemption_reason"] = ""
            gkey, glabel = _resolve_exemption_group(p, config)
            p["exemption_group"] = gkey
            p["exemption_group_label"] = glabel
        exempted_count = 0

    non_discontinue = [p for p in products if not p["discontinued"]]
    in_stock = [p for p in non_discontinue if p["in_stock"]]
    displayed_in_stock = [p for p in in_stock if p["displayed"]]
    not_displayed = [p for p in in_stock if p["gap"]]
    if store_specific:
        raw_gap_active = [
            p for p in products
            if p["in_stock"] and not p["discontinued"] and not p["displayed"]
        ]
        in_stock_not_displayed_all = [
            p for p in products
            if p["in_stock"] and not p["displayed"] and not p.get("exempted")
        ]
        in_stock_not_displayed_discontinued = [
            p for p in in_stock_not_displayed_all if p["discontinued"]
        ]
    else:
        raw_gap_active = []
        in_stock_not_displayed_all = []
        in_stock_not_displayed_discontinued = []

    total_nd = len(non_discontinue)
    in_stock_n = len(in_stock)
    summary = {
        "store": store,
        "region": region_key,
        "store_specific": store_specific,
        "total_non_discontinue": total_nd,
        "in_stock_count": in_stock_n,
        "in_stock_rate": round(in_stock_n / total_nd * 100, 2) if total_nd else None,
        "displayed_in_stock_count": len(displayed_in_stock),
        "display_coverage_rate": round(len(displayed_in_stock) / in_stock_n * 100, 2) if in_stock_n else None,
        "raw_gap_active_count": len(raw_gap_active) if store_specific else None,
        "not_displayed_count": len(not_displayed) if store_specific else None,
        "exempted_count": exempted_count if store_specific else None,
        "in_stock_not_displayed_all": len(in_stock_not_displayed_all) if store_specific else None,
        "in_stock_not_displayed_discontinued": len(in_stock_not_displayed_discontinued) if store_specific else None,
        "stock_sources": " + ".join(
            WAREHOUSE_LABELS.get(k, k) for k in _warehouses_for_store(store, region_key)
        ),
        "blacklist_count": len(blacklist),
    }

    products.sort(key=lambda p: (
        p.get("exemption_group_label") or p.get("family") or "未分类",
        not p["gap"],
        not p.get("exempted"),
        not p["in_stock"],
        -p["stock_qty"],
        p["code"],
    ))

    diagnostics = []
    display_row_count = bundle["display_row_count"]
    if display_row_count < 5:
        diagnostics.append({
            "level": "warning",
            "message": (
                f"展示数据几乎为空（display 仅 {display_row_count} 行）："
                f"{display_path}。请检查 Data-NZ/display_with_families.sql 并重新执行导出。"
            ),
        })
    elif len(by_store) == 0:
        diagnostics.append({
            "level": "warning",
            "message": (
                "展示数据里没有识别到店面列（需要 Store / DisplayWarehouse / Warehouse 等列名）。"
                f"当前文件：{display_path}"
            ),
        })
    elif len(stores) <= 1:
        diagnostics.append({
            "level": "warning",
            "message": "展示数据未包含有效店面，店面下拉只会显示「全部店面」。",
        })

    result = {
        "source": source,
        "region": region_key,
        "stock_path": str(stock_path),
        "display_path": str(display_path),
        "blacklist_path": blacklist_path,
        "blacklist_count": len(blacklist),
        "display_row_count": display_row_count,
        "stores": stores,
        "selected_store": store,
        "summary": summary,
        "products": products,
        "regions": list_regions(),
        "diagnostics": diagnostics,
    }
    _STORE_VIEW_CACHE[view_key] = result
    if only_gap:
        out = dict(result)
        out["products"] = [p for p in products if p["gap"]]
        return out
    return result


if __name__ == "__main__":
    import json as _json
    data = build_products(store=ALL_STORES, region=default_region())
    print("stores:", data["stores"][:5], "...")
    print("summary:", _json.dumps(data["summary"], ensure_ascii=False))
